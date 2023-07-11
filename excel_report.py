import pandas as pd
import numpy as np
from xlsxwriter.utility import xl_rowcol_to_cell
import locale
from openpyxl import load_workbook
from openpyxl.chart import reference, bar_chart
import string

locale.setlocale(locale.LC_ALL,'Portuguese')


#  df = Data Frame
df = pd.read_excel("supermarket_sales.xlsx")


df = df[['Gender','Product line','Unit price','Quantity','Total']]  # Seleciono as Colunas

# Rename Columns
#df.rename(columns={'Gender':'Genero','Product line':'Produtos','Unit price':'Preço Unitario','Quantity':'Quantidade'})

df = df[(df['Product line'] == 'Electronic accessories')]  # Filtrei o conteudo da coluna Produtos

df.to_excel('Eletronicos.xlsx',index=False,sheet_name='Relatorio')  # Exportei o Relatorio como Excel

# Summary
Resumo = pd.pivot_table(
    data=df,
    index='Gender',
    values=['Unit price','Quantity','Total'],
    aggfunc='sum'
)

Resumo.to_excel('Resumo.xlsx')

wb = load_workbook('Resumo.xlsx')
ws = wb['Sheet1']

# Adding Headers
ws.insert_rows(0,3)
ws['A1'].value = 'Vendas por Genero'
ws['A2'].value = 'Linha de Eletronicos'

ws['A1'].style = 'Title'
ws['A2'].style = 'Headline 2'

MaxColuna = ws.max_column
Alfabeto = list(string.ascii_uppercase)
Alfabeto_Excel = Alfabeto[0:MaxColuna]

for i in Alfabeto_Excel:
    if i != 'A' and i != 'B':
        for cell in range(5, ws.max_row+1):
            ws[f'{i}{cell}'].style = 'Currency'


wb.save('Resumo.xlsx')
